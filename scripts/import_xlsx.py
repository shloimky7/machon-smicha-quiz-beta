#!/usr/bin/env python3
import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


HEADERS = {
    "question": "Question",
    "option_a": "Option A",
    "option_b": "Option B",
    "option_c": "Option C",
    "option_d": "Option D",
    "correct_option": "Correct Option",
    "difficulty": "Difficulty Level",
    "siman_seif": "Siman and Seif",
    "page_numbers": "Page Number(s)",
    "correct_explanation": "Explanation Why the Correct Answer Is Correct",
    "why_a": "Why Option A Is Not Correct",
    "why_b": "Why Option B Is Not Correct",
    "why_c": "Why Option C Is Not Correct",
    "why_d": "Why Option D Is Not Correct",
}


def slugify(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "quiz"


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def default_target_count(bank_size: int) -> int:
    if bank_size >= 60:
        return 30
    return max(1, round(bank_size / 2))


def read_quiz(args):
    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Workbook has no rows.")

    header_row = [cell_text(v) for v in rows[0]]
    index = {name: header_row.index(label) for name, label in HEADERS.items() if label in header_row}
    missing = [label for key, label in HEADERS.items() if key not in index]
    if missing:
        raise SystemExit(f"Missing expected columns: {', '.join(missing)}")

    questions = []
    for row_number, row in enumerate(rows[1:], start=2):
        question = cell_text(row[index["question"]])
        if not question:
            continue
        qid = f"{args.id}-q{len(questions)+1:03d}"
        questions.append(
            {
                "id": qid,
                "question": question,
                "options": {
                    "A": cell_text(row[index["option_a"]]),
                    "B": cell_text(row[index["option_b"]]),
                    "C": cell_text(row[index["option_c"]]),
                    "D": cell_text(row[index["option_d"]]),
                },
                "correctOption": cell_text(row[index["correct_option"]]).upper(),
                "difficulty": cell_text(row[index["difficulty"]]),
                "source": {
                    "simanSeif": cell_text(row[index["siman_seif"]]),
                    "pageNumbers": cell_text(row[index["page_numbers"]]),
                },
                "explanations": {
                    "correct": cell_text(row[index["correct_explanation"]]),
                    "A": cell_text(row[index["why_a"]]),
                    "B": cell_text(row[index["why_b"]]),
                    "C": cell_text(row[index["why_c"]]),
                    "D": cell_text(row[index["why_d"]]),
                },
            }
        )

    return {
        "id": args.id,
        "title": args.title,
        "course": args.course,
        "kind": args.kind,
        "targetQuestionCount": args.target_count or default_target_count(len(questions)),
        "difficultyPolicy": {"Easy": 0.5, "Medium": 0.3, "Hard": 0.2},
        "sourceWorkbook": Path(args.xlsx).name,
        "questions": questions,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx")
    parser.add_argument("--id", required=True)
    parser.add_argument("--title", required=True)
    parser.add_argument("--course", default="Machon Smicha")
    parser.add_argument("--kind", default="quiz")
    parser.add_argument("--target-count", type=int)
    parser.add_argument("--sheet")
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    quiz = read_quiz(args)
    output = Path(args.out)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(quiz, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
